import re
import pyexcel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
import requests
import os, sys, shlex
import tkinter as tk
from tkinter import ttk

class Inventory():
    def __init__(self, root):

        self.frame = ttk.Frame(root)
        self.frame.pack(padx=10, pady=10, fill='x', expand=True)

        self.submit_button = ttk.Button(
           self.frame,
           text="Open inventory",
           command = get_inventory_from_site
        )
        self.submit_button.pack()




def get_data(data):
    pattern = "<td>\d{1,3}<\/td><td>(iPhone)(\s)?(X?|XR?|XS?)?(\d{1,2})?(S)?(\+?)\s?(mini)?(Pro)?\s?(Max)?<\/td><td>LCD<\/td><td>(Black?|White)<\/td><td>AFTERMARKET<\/td><td>(\d{1,3})"
    filtered_list = []

    web_data = re.findall(pattern, data)
    for line in list(web_data):
        data = []
        name = ""
        name_need_spaces = ["Pro", "mini", "Max"]
        for i in range(9):
            if line[i] in name_need_spaces:
                name += " "
            name += line[i]

        data.append(name)
        data.append(line[9])
        data.append(line[10])

        filtered_list.append(data)
    return filtered_list



def get_inventory_from_site():

    URL = "http://www.254repair.com/repairs//inshop2/inventoryreportapple.php"
    page = requests.get(URL)


    workbook = Workbook()
    ws = workbook.active

    data = get_data(page.text)
    ws.column_dimensions['A'].width = 25
    ws.append(["Phone", "Color", "Quanity", "Front", "Back", "Diff"])
    for d in data:
        ws.append(d)


    row_count = ws.max_row

    thin = Side(border_style="thin", color="000000")
    yellow = "ADD8E6"
    for rows in ws.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=6):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    workbook.save("inventory.xlsx")

    path = os.path.join(os.getcwd(), "inventory.xlsx")
    print(f"path:{path}")
    os.system("open " + shlex.quote(path))
